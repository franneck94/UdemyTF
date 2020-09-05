import osrm
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from datetime import date, timedelta
import numpy as np

excel_file = "dortmund_straßennamen_location_route.xlsx"
excel_file_new = "dortmund_straßennamen_location_route_distanz.xlsx"

wb = load_workbook(filename=excel_file, read_only=True)
ws = wb.active
wb_new = Workbook()
ws_new = wb_new.active
rows = list(ws.rows)[1:]
num_rows = len(rows)

MyConfig = osrm.RequestConfig("localhost:5000/v1/car")

row = ["", "uhrzeit","Straße Start","Nr Start",
    "Stadt Start","Lat Start","Lon Start","Straße Ziel",
    "Nr Ziel","Stadt Ziel","Lat Ziel","Lon Ziel", 
    "OSRM Dauer", "OSRM Distanz"]
ws_new.append(row)

for i, row in enumerate(rows):
    print(i+1, "/", num_rows)
    row = [cell.value for cell in row]

    try:
        lat1 = float(row[5])
        long1 = float(row[6])
        lat2 = float(row[10])
        long2 = float(row[11])

        p1 = osrm.Point(
            latitude=lat1,
            longitude=long1)
        p2 = osrm.Point(
            latitude=lat2, 
            longitude=long2)

        result = osrm.simple_route(p1, p2, url_config=MyConfig, output='route', geometry='wkt')[0]

        row += [result['duration']/60.0, result['distance']/1000.0]
        ws_new.append(row)

    except Exception as e:
        print("Error: ", e)

wb_new.save(excel_file_new)